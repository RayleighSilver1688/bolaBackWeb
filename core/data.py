import csv
from pathlib import Path
from openpyxl import load_workbook

data_dir = Path(__file__).parent.parent / "data"


def read_csv(filename):
    path = data_dir / filename

    with open(path, encoding="utf-8") as f:
        reader = csv.DictReader(f)

        for i in reader:
            yield list(i.values())


def read_excel(filename):
    path = data_dir / filename
    wb = load_workbook(path)
    ws = wb.active

    yield from ws.iter_rows(
        min_row=2,
        values_only=True
    )


if __name__ == '__main__':
    for d in read_excel('test11.xlsx'):
        print(d)
